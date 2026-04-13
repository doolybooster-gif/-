from __future__ import annotations

import csv
import io
import os
import re
import sqlite3
import sys
import uuid
from contextlib import closing
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = BASE_DIR / "static" / "uploads"
DB_PATH = DATA_DIR / "app.db"

DATA_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="QR 차량 2부제 점검 시스템")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

PLATE_PATTERN = re.compile(r"(\d{2,3}[가-힣]\d{4})")
DEFAULT_BRANCHES = [
    ("HQ", "본부"),
    ("ICN", "인천지사"),
    ("SWN", "수원지사"),
    ("BSN", "부산지사"),
]

def ensure_sample_branch_file() -> None:
    sample_path = BASE_DIR / "sample_branch_list.csv"
    if sample_path.exists():
        return

    content = (
        "code,name\n"
        "HQ,본부\n"
        "ICN,인천지사\n"
        "SWN,수원지사\n"
        "BSN,부산지사\n"
    )
    sample_path.write_text(content, encoding="utf-8-sig")
    
def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with closing(get_conn()) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS branches (
                code TEXT PRIMARY KEY,
                name TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS vehicles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_code TEXT,
                plate_no TEXT NOT NULL UNIQUE,
                owner_name TEXT,
                department TEXT,
                is_target INTEGER NOT NULL DEFAULT 1,
                exempt INTEGER NOT NULL DEFAULT 0,
                note TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(branch_code) REFERENCES branches(code)
            );

            CREATE TABLE IF NOT EXISTS scans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                branch_code TEXT NOT NULL,
                inspector_name TEXT,
                image_path TEXT,
                ocr_plate TEXT,
                confirmed_plate TEXT,
                vehicle_found INTEGER NOT NULL DEFAULT 0,
                is_target INTEGER NOT NULL DEFAULT 0,
                exempt INTEGER NOT NULL DEFAULT 0,
                is_violation INTEGER NOT NULL DEFAULT 0,
                violation_registered INTEGER NOT NULL DEFAULT 0,
                scan_date TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(branch_code) REFERENCES branches(code)
            );
            """
        )
        for code, name in DEFAULT_BRANCHES:
            conn.execute(
                "INSERT OR IGNORE INTO branches(code, name) VALUES(?, ?)",
                (code, name),
            )
        conn.commit()


init_db()
ensure_sample_branch_file()

def today_allowed_digit_parity(target_date: date) -> int:
    """짝수일 0, 홀수일 1. 오늘 날짜와 같은 홀짝 차량이 운행 가능."""
    return target_date.day % 2

def is_violation_plate(plate_no: str, target_date: date) -> bool:
    digits = re.findall(r"\d", plate_no)
    if not digits:
        return False

    last_digit = int(digits[-1])
    return last_digit % 2 != today_allowed_digit_parity(target_date)


# Optional OCR support.
def run_ocr(image_bytes: bytes) -> Optional[str]:
    try:
        import pytesseract  # type: ignore
        from PIL import Image  # type: ignore
    except Exception:
        return None

    try:
        image = Image.open(io.BytesIO(image_bytes))
        raw = pytesseract.image_to_string(image, lang="kor+eng", config="--psm 7")
        compact = re.sub(r"\s+", "", raw)
        match = PLATE_PATTERN.search(compact)
        return match.group(1) if match else None
    except Exception:
        return None

def is_full_plate(text: str) -> bool:
    text = re.sub(r"\s+", "", text).upper()
    return bool(PLATE_PATTERN.fullmatch(text))


def find_plate_candidates(branch_code: str, text: str) -> list[sqlite3.Row]:
    text = re.sub(r"\s+", "", text).upper()

    with closing(get_conn()) as conn:
        # 1) 완전한 차량번호 형식이면 정확히 일치 검색
        if is_full_plate(text):
            row = conn.execute(
                "SELECT * FROM vehicles WHERE plate_no = ?",
                (text,),
            ).fetchone()
            return [row] if row else []

        # 2) 마지막 4자리만 입력한 경우, 해당 지사 기준 후보 검색
        if re.fullmatch(r"\d{4}", text):
            rows = conn.execute(
                """
                SELECT *
                FROM vehicles
                WHERE branch_code = ?
                  AND substr(plate_no, -4) = ?
                ORDER BY plate_no
                """,
                (branch_code, text),
            ).fetchall()
            return list(rows)

    return []


@app.get("/", response_class=HTMLResponse)
def home(request: Request) -> Any:
    with closing(get_conn()) as conn:
        branches = conn.execute("SELECT code, name FROM branches ORDER BY name").fetchall()
        vehicle_count = conn.execute("SELECT COUNT(*) AS c FROM vehicles").fetchone()["c"]
        scan_count = conn.execute("SELECT COUNT(*) AS c FROM scans").fetchone()["c"]
        violation_count = conn.execute(
            "SELECT COUNT(*) AS c FROM scans WHERE violation_registered = 1"
        ).fetchone()["c"]
    return templates.TemplateResponse(
        request,
        "home.html",
        {
            "branches": branches,
            "vehicle_count": vehicle_count,
            "scan_count": scan_count,
            "violation_count": violation_count,
            "today": date.today().isoformat(),
            "allowed_parity": "짝수" if today_allowed_digit_parity(date.today()) == 0 else "홀수",
        },
    )


@app.get("/admin/vehicles", response_class=HTMLResponse)
def admin_vehicles(request: Request) -> Any:
    with closing(get_conn()) as conn:
        vehicles = conn.execute(
            """
            SELECT v.*, b.name AS branch_name
            FROM vehicles v
            LEFT JOIN branches b ON b.code = v.branch_code
            ORDER BY v.branch_code, v.plate_no
            LIMIT 200
            """
        ).fetchall()
        branches = conn.execute("SELECT code, name FROM branches ORDER BY name").fetchall()
    return templates.TemplateResponse(
        request,
        "admin_vehicles.html",
        {"vehicles": vehicles, "branches": branches, "sample_file": "/sample_vehicle_list.csv"},
    )

@app.get("/admin/branches", response_class=HTMLResponse)
def admin_branches(request: Request) -> Any:
    with closing(get_conn()) as conn:
        branches = conn.execute(
            "SELECT code, name FROM branches ORDER BY name"
        ).fetchall()

    return templates.TemplateResponse(
        request,
        "admin_branches.html",
        {
            "branches": branches,
            "sample_file": "/sample_branch_list.csv",
        },
    )

@app.post("/admin/vehicles/upload")
async def upload_vehicle_list(file: UploadFile = File(...)) -> RedirectResponse:
    suffix = Path(file.filename or "").suffix.lower()
    raw = await file.read()
    rows: list[dict[str, Any]] = []

    if suffix == ".csv":
        decoded = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    elif suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {headers[i]: row[i] for i in range(len(headers))}
            rows.append(item)
    else:
        raise HTTPException(status_code=400, detail="CSV 또는 XLSX 파일만 업로드할 수 있습니다.")

    now = datetime.now().isoformat(timespec="seconds")
    inserted = 0
    with closing(get_conn()) as conn:
        for row in rows:
            plate_no = str(row.get("plate_no") or row.get("차량번호") or "").strip().upper()
            if not plate_no:
                continue
            branch_code = str(row.get("branch_code") or row.get("지사코드") or "").strip() or None
            owner_name = str(row.get("owner_name") or row.get("성명") or "").strip() or None
            department = str(row.get("department") or row.get("부서") or "").strip() or None
            is_target = normalize_bool(row.get("is_target") or row.get("2부제대상") or "Y")
            exempt = normalize_bool(row.get("exempt") or row.get("예외") or "N")
            note = str(row.get("note") or row.get("비고") or "").strip() or None

            conn.execute(
                """
                INSERT INTO vehicles(branch_code, plate_no, owner_name, department, is_target, exempt, note, created_at, updated_at)
                VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(plate_no) DO UPDATE SET
                    branch_code=excluded.branch_code,
                    owner_name=excluded.owner_name,
                    department=excluded.department,
                    is_target=excluded.is_target,
                    exempt=excluded.exempt,
                    note=excluded.note,
                    updated_at=excluded.updated_at
                """,
                (branch_code, plate_no, owner_name, department, is_target, exempt, note, now, now),
            )
            inserted += 1
        conn.commit()
    return RedirectResponse(url=f"/admin/vehicles?uploaded={inserted}", status_code=303)

@app.post("/admin/branches/upload")
async def upload_branch_list(file: UploadFile = File(...)) -> RedirectResponse:
    suffix = Path(file.filename or "").suffix.lower()
    raw = await file.read()
    rows: list[dict[str, Any]] = []

    if suffix == ".csv":
        decoded = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        rows = list(reader)
    elif suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {headers[i]: row[i] for i in range(len(headers))}
            rows.append(item)
    else:
        raise HTTPException(status_code=400, detail="CSV 또는 XLSX 파일만 업로드할 수 있습니다.")

    inserted = 0
    with closing(get_conn()) as conn:
        for row in rows:
            code = str(row.get("code") or row.get("지사코드") or "").strip().upper()
            name = str(row.get("name") or row.get("지사명") or "").strip()

            if not code or not name:
                continue

            conn.execute(
                """
                INSERT INTO branches(code, name)
                VALUES(?, ?)
                ON CONFLICT(code) DO UPDATE SET
                    name=excluded.name
                """,
                (code, name),
            )
            inserted += 1

        conn.commit()

    return RedirectResponse(url=f"/admin/branches?uploaded={inserted}", status_code=303)

@app.get("/sample_vehicle_list.csv")
def sample_vehicle_list() -> StreamingResponse:
    content = (
        "branch_code,plate_no,owner_name,department,is_target,exempt,note\n"
        "ICN,12가3456,홍길동,총무팀,Y,N,상시점검\n"
        "ICN,234나5678,김영수,시설팀,Y,N,\n"
        "SWN,123다4567,박민수,행정지원부,Y,Y,공무수행 예외\n"
    )
    return StreamingResponse(iter([content.encode("utf-8-sig")]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=sample_vehicle_list.csv"})

@app.get("/sample_branch_list.csv")
def sample_branch_list() -> StreamingResponse:
    content = (
        "code,name\n"
        "HQ,본부\n"
        "ICN,인천지사\n"
        "SWN,수원지사\n"
        "BSN,부산지사\n"
    )
    return StreamingResponse(
        iter([content.encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sample_branch_list.csv"},
    )

@app.get("/branch/{branch_code}", response_class=HTMLResponse)
def branch_scan_page(request: Request, branch_code: str) -> Any:
    with closing(get_conn()) as conn:
        branch = conn.execute("SELECT * FROM branches WHERE code = ?", (branch_code,)).fetchone()
        if not branch:
            raise HTTPException(status_code=404, detail="지사를 찾을 수 없습니다.")
    return templates.TemplateResponse(
        request,
        "scan.html",
        {"branch": branch, "today": date.today().isoformat(), "allowed_parity": "짝수" if today_allowed_digit_parity(date.today()) == 0 else "홀수"},
    )


@app.post("/api/scan")
async def api_scan(
    branch_code: str = Form(...),
    inspector_name: str = Form(""),
    manual_plate: str = Form(""),
    image: UploadFile = File(...),
) -> dict[str, Any]:
    with closing(get_conn()) as conn:
        branch = conn.execute("SELECT * FROM branches WHERE code = ?", (branch_code,)).fetchone()
        if not branch:
            raise HTTPException(status_code=404, detail="지사를 찾을 수 없습니다.")

    payload = await image.read()
    ext = Path(image.filename or "plate.jpg").suffix or ".jpg"
    file_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    target_path = UPLOAD_DIR / file_name
    target_path.write_bytes(payload)

    ocr_plate = run_ocr(payload)
    manual_text = re.sub(r"\s+", "", manual_plate).upper()

    candidate_plates: list[str] = []
    confirmed_plate = ""

if manual_text:
    candidates = find_plate_candidates(branch_code, manual_text)

    # 1건이면 자동 확정
    if len(candidates) == 1:
        confirmed_plate = candidates[0]["plate_no"]

    # 여러 건이면 후보 선택 필요
    elif len(candidates) > 1:
        candidate_plates = [c["plate_no"] for c in candidates]

    # 후보가 없으면, 완전한 번호판 형식일 때만 그대로 사용
    else:
        confirmed_plate = manual_text if is_full_plate(manual_text) else ""
else:
    confirmed_plate = (ocr_plate or "").strip().upper()

    if candidate_plates:
    return {
        "scan_id": None,
        "ocr_plate": ocr_plate,
        "confirmed_plate": None,
        "vehicle_found": False,
        "vehicle": None,
        "is_violation": False,
        "needs_selection": True,
        "candidates": candidate_plates,
        "message": "끝 4자리가 일치하는 차량이 여러 대입니다. 차량번호를 선택해 주세요.",
    }

if manual_text and not confirmed_plate:
    return {
        "scan_id": None,
        "ocr_plate": ocr_plate,
        "confirmed_plate": None,
        "vehicle_found": False,
        "vehicle": None,
        "is_violation": False,
        "needs_selection": False,
        "candidates": [],
        "message": "입력한 번호와 일치하는 차량을 찾지 못했습니다. 차량번호 전체를 입력하거나 다시 확인해 주세요.",
    }
    
    vehicle = None
    if confirmed_plate:
        with closing(get_conn()) as conn:
            vehicle = conn.execute("SELECT * FROM vehicles WHERE plate_no = ?", (confirmed_plate,)).fetchone()

    scan_date = date.today()
    vehicle_found = 1 if vehicle else 0
    is_target = int(vehicle["is_target"]) if vehicle else 0
    exempt = int(vehicle["exempt"]) if vehicle else 0
    is_violation = int(bool(vehicle and is_target and not exempt and is_violation_plate(confirmed_plate, scan_date)))

    now = datetime.now().isoformat(timespec="seconds")
    with closing(get_conn()) as conn:
        cur = conn.execute(
            """
            INSERT INTO scans(
                branch_code, inspector_name, image_path, ocr_plate, confirmed_plate,
                vehicle_found, is_target, exempt, is_violation, violation_registered,
                scan_date, created_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
            """,
            (
                branch_code,
                inspector_name.strip() or None,
                f"/static/uploads/{file_name}",
                ocr_plate,
                confirmed_plate or None,
                vehicle_found,
                is_target,
                exempt,
                is_violation,
                scan_date.isoformat(),
                now,
            ),
        )
        scan_id = cur.lastrowid
        conn.commit()

    return {
        "scan_id": scan_id,
        "ocr_plate": ocr_plate,
        "confirmed_plate": confirmed_plate,
        "vehicle_found": bool(vehicle_found),
        "vehicle": {
            "plate_no": vehicle["plate_no"],
            "owner_name": vehicle["owner_name"],
            "department": vehicle["department"],
            "is_target": bool(vehicle["is_target"]),
            "exempt": bool(vehicle["exempt"]),
            "note": vehicle["note"],
        } if vehicle else None,
        "is_violation": bool(is_violation),
        "message": make_scan_message(vehicle, confirmed_plate, bool(is_violation)),
    }


@app.post("/api/scan/{scan_id}/register_violation")
def register_violation(scan_id: int, register: bool = Form(...)) -> dict[str, Any]:
    with closing(get_conn()) as conn:
        scan = conn.execute("SELECT * FROM scans WHERE id = ?", (scan_id,)).fetchone()
        if not scan:
            raise HTTPException(status_code=404, detail="점검 이력을 찾을 수 없습니다.")
        conn.execute(
            "UPDATE scans SET violation_registered = ? WHERE id = ?",
            (1 if register else 0, scan_id),
        )
        conn.commit()
    return {"ok": True, "registered": bool(register)}


@app.get("/admin/scans", response_class=HTMLResponse)
def admin_scans(request: Request) -> Any:
    with closing(get_conn()) as conn:
        scans = conn.execute(
            """
            SELECT s.*, b.name AS branch_name
            FROM scans s
            LEFT JOIN branches b ON b.code = s.branch_code
            ORDER BY s.id DESC
            LIMIT 200
            """
        ).fetchall()
    return templates.TemplateResponse(request, "admin_scans.html", {"scans": scans})


@app.get("/admin/export")
def export_excel() -> FileResponse:
    export_path = DATA_DIR / f"vehicle_ban_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    build_report_workbook(export_path)
    return FileResponse(path=export_path, filename=export_path.name)


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}



def normalize_bool(value: Any) -> int:
    text = str(value).strip().upper()
    return 1 if text in {"Y", "YES", "TRUE", "1", "대상", "예", "O"} else 0



def make_scan_message(vehicle: Optional[sqlite3.Row], confirmed_plate: str, is_violation: bool) -> str:
    if not confirmed_plate:
        return "차량번호를 인식하지 못했습니다. 수기로 보정해 주세요."
    if vehicle is None:
        return f"{confirmed_plate} 차량은 사전 등록 리스트에 없습니다."
    if vehicle["exempt"]:
        return f"{confirmed_plate} 차량은 예외 차량으로 등록되어 있습니다."
    if not vehicle["is_target"]:
        return f"{confirmed_plate} 차량은 2부제 적용 대상이 아닙니다."
    if is_violation:
        return f"{confirmed_plate} 차량은 오늘 2부제 위반 대상입니다. 위반 등록 여부를 선택해 주세요."
    return f"{confirmed_plate} 차량은 등록 차량이지만 오늘 위반 대상은 아닙니다."



def build_report_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "점검요약"

    title_fill = PatternFill("solid", fgColor="D9EAD3")
    section_fill = PatternFill("solid", fgColor="DDEBF7")
    header_fill = PatternFill("solid", fgColor="FCE4D6")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    with closing(get_conn()) as conn:
        summary_rows = conn.execute(
            """
            SELECT b.name AS branch_name,
                   COUNT(s.id) AS total_scans,
                   SUM(CASE WHEN s.vehicle_found = 1 THEN 1 ELSE 0 END) AS registered_hits,
                   SUM(CASE WHEN s.is_violation = 1 THEN 1 ELSE 0 END) AS violation_targets,
                   SUM(CASE WHEN s.violation_registered = 1 THEN 1 ELSE 0 END) AS violation_registered
            FROM branches b
            LEFT JOIN scans s ON s.branch_code = b.code
            GROUP BY b.code, b.name
            ORDER BY b.name
            """
        ).fetchall()
        detailed_rows = conn.execute(
            """
            SELECT s.scan_date, b.name AS branch_name, s.inspector_name, s.ocr_plate, s.confirmed_plate,
                   s.vehicle_found, s.is_target, s.exempt, s.is_violation, s.violation_registered, s.image_path, s.created_at
            FROM scans s
            LEFT JOIN branches b ON b.code = s.branch_code
            ORDER BY s.id DESC
            """
        ).fetchall()

    ws.merge_cells("A1:H1")
    ws["A1"] = "차량 2부제 점검 현황"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center

    ws["A3"] = "1. 지사별 점검 요약"
    ws["A3"].font = bold
    ws["A3"].fill = section_fill

    headers = ["지사", "점검건수", "등록차량 일치", "위반대상 판정", "최종 위반등록"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    row_no = 5
    for row in summary_rows:
        ws.cell(row=row_no, column=1, value=row["branch_name"])
        ws.cell(row=row_no, column=2, value=row["total_scans"] or 0)
        ws.cell(row=row_no, column=3, value=row["registered_hits"] or 0)
        ws.cell(row=row_no, column=4, value=row["violation_targets"] or 0)
        ws.cell(row=row_no, column=5, value=row["violation_registered"] or 0)
        row_no += 1

    detail = wb.create_sheet("위반차량상세")
    detail_headers = [
        "점검일", "지사", "점검자", "OCR 인식", "확정 차량번호", "등록차량 여부",
        "2부제 대상", "예외", "위반판정", "위반등록", "사진경로", "등록시각"
    ]
    for idx, header in enumerate(detail_headers, start=1):
        cell = detail.cell(row=1, column=idx, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    for r, item in enumerate(detailed_rows, start=2):
        detail.cell(r, 1, item["scan_date"])
        detail.cell(r, 2, item["branch_name"])
        detail.cell(r, 3, item["inspector_name"])
        detail.cell(r, 4, item["ocr_plate"])
        detail.cell(r, 5, item["confirmed_plate"])
        detail.cell(r, 6, "Y" if item["vehicle_found"] else "N")
        detail.cell(r, 7, "Y" if item["is_target"] else "N")
        detail.cell(r, 8, "Y" if item["exempt"] else "N")
        detail.cell(r, 9, "Y" if item["is_violation"] else "N")
        detail.cell(r, 10, "Y" if item["violation_registered"] else "N")
        detail.cell(r, 11, item["image_path"])
        detail.cell(r, 12, item["created_at"])

    raw = wb.create_sheet("차량기준정보")
    raw_headers = ["지사코드", "차량번호", "성명", "부서", "2부제대상", "예외", "비고"]
    for idx, header in enumerate(raw_headers, start=1):
        cell = raw.cell(row=1, column=idx, value=header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    with closing(get_conn()) as conn:
        vehicle_rows = conn.execute(
            "SELECT branch_code, plate_no, owner_name, department, is_target, exempt, note FROM vehicles ORDER BY branch_code, plate_no"
        ).fetchall()
    for r, item in enumerate(vehicle_rows, start=2):
        raw.cell(r, 1, item["branch_code"])
        raw.cell(r, 2, item["plate_no"])
        raw.cell(r, 3, item["owner_name"])
        raw.cell(r, 4, item["department"])
        raw.cell(r, 5, "Y" if item["is_target"] else "N")
        raw.cell(r, 6, "Y" if item["exempt"] else "N")
        raw.cell(r, 7, item["note"])

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18

    wb.save(path)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)
